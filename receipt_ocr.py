import re, os, json
import cv2
import numpy as np
import openpyxl
import pandas as pd
import pytesseract
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
from dataclasses import dataclass, field

load_dotenv()
excel_path = os.environ["EXCEL_PATH"]
STATE_JSON = os.environ.get("STATE_JSON")

PLACEHOLDERS = ['REPLACE', 'Bot_Holder']

DATE_PATTERNS: tuple[re.Pattern,...] = (
    re.compile(r"\d{2}-[A-Z]{3}-\d{4}", re.I),
    re.compile(r"\d{2}-\d{2}-\d{4}"),
    re.compile(r"\d{2}-\d{2}-\d{2}"),
    re.compile(r"\d{2}/\d{2}/\d{4}"),
    re.compile(r"\d{2}/\d{2}/\d{2}"),
)

def gather_picture_files(dir_path: Path) -> list:
    receipt_list = []
    new_files = []
    files_read = None

    # Looking for JSON file
    try:
        with open(STATE_JSON) as f:
            files_read = json.load(f)["files_read"]
    except FileNotFoundError:
        print(f"[ERROR] File {STATE_JSON} not found.")
    except KeyError as e:
        print(f"[ERROR] KEY {e} not found.")

    # Checking values returned
    if files_read:
        # Searching for files not in files_read
        for file in dir_path.iterdir():
            if str(file) not in files_read:
                new_files.append(file)

        return new_files

    # no values in JSON
    else:
        for file in dir_path.iterdir():
            receipt_list.append(file)
        return receipt_list

@dataclass(slots=False)
class ReceiptOCR:
    receipt_path: Path
    config: Optional[str] = field(default=None)

    purchase_date: Optional[str] = field(init=False, default=None)
    supplier: Optional[str] = field(init=False, default=None)
    cost_text: Optional[float] = field(init=False, default=None)
    text: Optional[str] = field(init=False, repr=False, default=None)
    data: Optional[str] = field(init=False, repr=False, default=None)

    def __post_init__(self) -> None:
        try:
            image = cv2.imread(str(self.receipt_path))
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

            if self.config:
                self.text = pytesseract.image_to_string(image, config=self.config)
                self.data = pytesseract.image_to_string(image, config=self.config)

            # This is the default config case
            else:
                self.text = pytesseract.image_to_string(image)
                self.data = pytesseract.image_to_string(image, config=self.config)

            self._extract_text(self.text)

        except Exception as e:
            print(f"[ERROR] {e}")

    def _text_analysis (self, data: pd.DataFrame):
        mask = self.data
        n_rows = mask.sum()
        mask_70 = (self.data["conf"] >=70)
        stats = data.loc[mask,'conf'].agg(['count','mean','std'])
        bins = pd.cut(mask['conf'], [-1, 0, 50, 70, 85, 100], labels=['<0', '0–50', '50–70', '70–85', '85–100'])
        dist_70 = bins.value_counts().loc['85–100']
        perc = dist_70.sum()/n_rows

        print(mask_70)
        print(stats)
        print(perc)

    def _extract_text(self, text: str) -> None:
        print("entered")

        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        if lines:
            self.supplier = lines[0]

        for line in lines:
            if "total" in line.lower():
                amount_pattern = r"\$?\s?\d[\d,]*\.?\d{0,2}"
                match = re.search(amount_pattern, line)
                print(match)

                if match:
                    self.cost_text = match.group(0).replace(" ", "")
                    break

        for line in lines:
            for pattern in DATE_PATTERNS:
                match = pattern.search(line)
                if match:
                    self.purchase_date = match.group(0)
                    break

            if self.purchase_date:
                break

        self._text_analysis()

def process_receipt (receipt: Path) ->ReceiptOCR:
    return ReceiptOCR(receipt_path=receipt)

def upload_file_tracking(receipt: Path):
    try:
        with open(STATE_JSON) as f:
            files_read = json.load(f)['files_read']
    except FileNotFoundError:
        files_read = []

    files_read.append(str(receipt))

    with open(STATE_JSON, 'w') as f:
        json.dump({'files_read': files_read}, f,indent=2)

def extracted_text_to_excel(text: pd.DataFrame):
    # will move this to the main app.py
    # excel_path = os.environ["EXCEL_PATH"]

    wb = openpyxl.load_workbook(excel_path)
    count = len(wb.sheetnames)
    sheet_name = f"Sheet{count+1}"

    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")

    with pd.ExcelWriter(excel_path, mode='a' ,engine='openpyxl', if_sheet_exists='replace') as writer:
        text.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = openpyxl.styles.Font(color='9C0006')

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Bot_Holder":
                cell.fill = red_fill

            elif cell.value == "Null":
                cell.value = " "

            if cell.value == "REPLACE":
                cell.font = red_font
    wb.save(excel_path)

def combine_data_sources(ocr_data: dict):
    # reading excel
    excel = pd.read_excel(excel_path, sheet_name='Sheet1')
    df = pd.DataFrame.from_dict(ocr_data, orient='index')

    df_reset = df.reset_index().rename(columns={'index':'Receipt_Number'})
    col_names = df_reset.columns.drop('Receipt_Number')

    # Merging the data and adding suffix to separate
    merged_data = df_reset.merge(excel,on='Receipt_Number', suffixes=('_old', ''))
    for col in col_names:
        merged_data[f"{col}_old"] = merged_data[f"{col}_old"].replace(PLACEHOLDERS, np.nan)

    for col in col_names:
        merged_data[col] = merged_data[f'{col}_old'].fillna(merged_data[col])

    to_drop = merged_data.columns[merged_data.columns.str.endswith('_old')]
    merged_data.drop(to_drop, axis=1, inplace=True)

    merged_data = merged_data[["Download_Date", "Purchase_Name","Purchase_Date","Description",
        "Supplier", "Receipt_Number", "Cost", "Message", "Purchaser", "Reimbursed", "Error_Flag"]]

    extracted_text_to_excel(merged_data)

# Start of pipeline
def ocr_pipeline():
    download_dir = Path(os.environ['DOWNLOAD_LOC'])

    # Getting all receipts that exist within the Directory
    receipt_list = gather_picture_files(download_dir)

    # OCR output configuration
    results_def: dict[str, dict[str, Optional[str]]] = {}

    # Setting up OCR for each receipt
    for receipt in receipt_list:
        rec = process_receipt(receipt)
        print(receipt)
        print(rec)

        # update JSON files read
        upload_file_tracking(receipt)

        results_def[receipt.stem] = {
            "Purchase_Date": rec.purchase_date,
            "Supplier": rec.supplier,
            "Cost": rec.cost_text or None,
        }

    # quick debug print
    for k, v in results_def.items():
        print(k, "→", v)

# ocr_pipeline()