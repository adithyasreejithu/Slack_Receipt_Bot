import re
import cv2
import os, pathlib
import openpyxl
import pandas as pd
import numpy as np
import pytesseract
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from dataclasses import dataclass, field
from typing import Optional, Iterable


CONFIG1 = r"--oem 3 --psm 6"
CONFIG2 = r"--oem 1 --psm 6"

DATE_PATTERNS: tuple[re.Pattern,...] = (
    re.compile(r"\d{2}-[A-Z]{3}-\d{4}", re.I),
    re.compile(r"\d{2}-\d{2}-\d{4}"),
    re.compile(r"\d{2}-\d{2}-\d{2}"),
    re.compile(r"\d{2}/\d{2}/\d{4}"),
    re.compile(r"\d{2}/\d{2}/\d{2}"),
)

PLACEHOLDERS = ['REPLACE', 'Bot_Holder']

# This gets all the files that are in the directory
#       - Need to create json date to read last file read
#       - Check image file types as well
def gather_picture_files(dir: pathlib.Path):
    receipt_list = []

    for file in dir.iterdir():
        receipt_list.append(file)

    return receipt_list

@dataclass(slots=False)
class ReceiptOCR:
    receipt_path: Path
    config: str = CONFIG1
    
    purchase_date: Optional[str] = field(init=False, default=None)
    supplier: Optional[str] = field(init=False, default=None)
    cost_text : Optional[float] = field(init=False, default=None)
    text: Optional[str] = field(init=False,repr= False, default=None)

    def __post_init__(self) -> None:
        try:
            img = cv2.imread(str(self.receipt_path))
        except FileNotFoundError:
            print(f"File not found: {self.receipt_path}")

        processed_img = self._preprocess_grey(img)
        self.text = pytesseract.image_to_string(processed_img, config=self.config)
        self._extract_text(self.text)

    def _preprocess_grey(self, img: np.ndarray) -> np.ndarray:
        # return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        grey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Need to do more research on and find better methods of thresh holds
        den = cv2.medianBlur(grey, 3)
        _, th = cv2.threshold(den, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return th

    # Extracts the text values from the images
    def _extract_text(self, text: str) -> None:
        # clean the text - for line in text split and strip trail and return if not empty
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        if lines:
            self.supplier = lines[0]

        for line in lines:
            if "Total" in line.lower():
                amount_pattern = r"\$?\s?\d[\d,]*\.?\d{0,2}"
                match = re.search(amount_pattern, line)
                print(match)

                if match:
                    self.cost_text = match.group(0).replace(" ", "")
                    break

        for line in lines:
            for pattern in DATE_PATTERNS:
                match = pattern.search(line)
                if match:
                    self.purchase_date = match.group(0)
                    break

            if self.purchase_date:
                break

# wrapper to call dataclass easier
def process_receipt(receipt: Path, config: str = CONFIG1) -> ReceiptOCR:
    return ReceiptOCR(receipt_path=receipt, config=config)

def format_excel_output(excel_path, wb: openpyxl.workbook.Workbook, sheet):

    ws = wb[sheet]
    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = openpyxl.styles.Font(color='9C0006')

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Bot_Holder" :
                cell.fill = red_fill

            elif cell.value == "Null":
                cell.value =" "

            if cell.value == "REPLACE":
                cell.font = red_font
    wb.save(excel_path)

def extracted_text_to_excel(text: pd.DataFrame):
    # will move this to the main app.py
    excel_path = os.environ["EXCEL_PATH"]

    wb = openpyxl.load_workbook(excel_path)
    count = len(wb.sheetnames)
    sheet_name = f"Sheet{count+1}"

    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")

    with pd.ExcelWriter(excel_path, mode='a' ,engine='openpyxl', if_sheet_exists='replace') as writer:
        text.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = openpyxl.styles.Font(color='9C0006')

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Bot_Holder":
                cell.fill = red_fill

            elif cell.value == "Null":
                cell.value = " "

            if cell.value == "REPLACE":
                cell.font = red_font
    wb.save(excel_path)


def combine_data_sources(ocr_data: dict):
    load_dotenv()
    excel_path = os.environ["EXCEL_PATH"]

    excel = pd.read_excel(excel_path, sheet_name='Sheet1')
    df = pd.DataFrame.from_dict(ocr_data, orient='index')

    df_reset = df.reset_index().rename(columns={'index':'Receipt_Number'})
    col_names = df_reset.columns.drop('Receipt_Number')

    merged_data = df_reset.merge(excel,on='Receipt_Number', suffixes=('_old', ''))
    for col in col_names:
        merged_data[f"{col}_old"] = merged_data[f"{col}_old"].replace(PLACEHOLDERS, np.nan)

    for col in col_names:
        merged_data[col] = merged_data[f'{col}_old'].fillna(merged_data[col])

    to_drop = merged_data.columns[merged_data.columns.str.endswith('_old')]
    merged_data.drop(to_drop, axis=1, inplace=True)

    merged_data = merged_data[["Download_Date", "Purchase_Name","Purchase_Date","Description",
        "Supplier", "Receipt_Number", "Cost", "Message", "Purchaser", "Reimbursed", "Error_Flag"]]

    extracted_text_to_excel(merged_data)



if __name__ == "__main__":
    load_dotenv()
    download_dir = Path(os.environ["DOWNLOAD_LOC"])

    receipts = gather_picture_files(download_dir)
    results_con_1 : dict[str, dict[str, Optional[str]]] = {}
    results_con_2 : dict[str, dict[str, Optional[str]]] = {}

    for receipt in receipts:
        rec = process_receipt(receipt, CONFIG1)
        rec2 = process_receipt(receipt, CONFIG2)

        results_con_1[receipt.stem] = {
            "Purchase_Date": rec.purchase_date,
            "Supplier": rec.supplier,
            "Cost": rec.cost_text or None,
        }

        results_con_2[receipt.stem] = {
            "Purchase_Date": rec2.purchase_date,
            "Supplier": rec2.supplier,
            "Cost": rec2.cost_text or None,
        }

    # right now only processes one but will need to look at results and find best one
    combine_data_sources(results_con_1)

        # quick debug print
    for k, v in results_con_1.items():
        print(k, "→", v)

    print("\n")

    for k, v in results_con_2.items():
        print(k, "→", v)



