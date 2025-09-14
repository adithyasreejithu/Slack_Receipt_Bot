import os, pathlib, requests, json
import pandas as pd 
from datetime import datetime
from dotenv import load_dotenv
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
import openpyxl
import pandas

load_dotenv()   # This loads the env variables 
BOT_TOKEN = os.environ["SLACK_BOT_TOKEN"]   # API key for bot
CHANNEL_ID = os.environ["CHANNEL_ID"] # Channel ID 
download_loc = os.environ["DOWNLOAD_LOC"]
JSON_FILE = os.environ["JSON_FILE"]
client = WebClient(token=BOT_TOKEN)
DOWNLOAD_DIR = pathlib.Path(download_loc)
DOWNLOAD_DIR.mkdir(exist_ok=True)
COLUMNS = ["Download_Date", "Purchase_Name","Purchase_Date","Description", "Supplier", "Cost", "Message", "Purchaser","Receipt_Number", "Reimbursed", "Error_Flag"]
    
# Call function to map users ID to name 
def create_user_map() -> dict[str, str]: 
    cursor = None
    user_map = {}
    while True: 
        try:
            response = client.users_list(limit=200, cursor=cursor)
            members = response.get("members",[])
            for m in members:
                user_id = m.get("id")
                profile = m.get("profile",{}) or {}
                name  = profile.get("real_name") or profile.get("display_name")
                user_map[user_id] = name

                
        except SlackApiError as e:
            print("users.list error:", e.response.get("error"))
            break

        cursor = response.get("response_metadata", {}).get("next_cursor")

        if not cursor:
            break

    return user_map

def tracking_generator(prefix="R"):
    counter = 0
    def tracking_number():
        nonlocal counter
        counter += 1 
        return f"{prefix}{counter:03d}"
    return tracking_number

# Need to rename the file
def change_file_name(file_path: pathlib.Path, directory: pathlib.Path, make_invoice): 
    directory.mkdir(parents=True, exist_ok=True)
    while True: 
        receipt_number = make_invoice()
        file = directory/f"{receipt_number}{file_path.suffix}"

        if not file.exists():
            break

    file_path.rename(file)
    return receipt_number, file

# Create funcition to download files 
def download_files(file_url : str, save_path : str):  # url_private_download from slack files json, and path to be saved
    req = requests.get(
        file_url,
        headers={"Authorization": f"Bearer {BOT_TOKEN}"},
        timeout= 30)
    
    req.raise_for_status()
    with open(save_path, "wb") as f:
        f.write(req.content)

def upload_collection_excel_local (info: dict):
    df = pd.DataFrame(data=info, index=[0])

    excel_path = os.environ["EXCEL_PATH"]

    if not os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, mode='w', engine='openpyxl') as writer:
            pd.DataFrame(columns=COLUMNS).to_excel(writer, sheet_name='Sheet1', index=False)

    with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)

# This can also be put in the main file as it can be reused
def format_excel_output(): 
    excel_path = os.environ["EXCEL_PATH"]

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet1"]

    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') 
    red_font = openpyxl.styles.Font(color='9C0006') 

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Bot_Holder" :
                cell.fill = red_fill

            elif cell.value == "Null": 
                cell.value =" "

            if cell.value == "REPLACE":
                cell.font = red_font
    wb.save(excel_path)

# Pipeline to sharepoint???

def channel_history(chan_id : str):
    try: 
        with open(JSON_FILE) as f :
            last_ts  = json.load(f)["last_ts"]
    except FileNotFoundError:
        last_ts = 0

    user_map = create_user_map()
    cursor = None 
    all_files = []
    rows = []

    column  =  ["Download_Date", "Purchase_Name","Purchase_Date","Description", "Supplier", "Cost", "Message", "Purchaser","Receipt_Number", "Reimbursed", "Error_Flag"]
    receipt_info = pd.DataFrame(columns=column)
    # receipt_info["Reimbursed"] = receipt_info["Reimbursed"].fillna("No")

    make_invoice = tracking_generator("R")

    defaults = {
        "Download_Date" : "Bot_Holder",  # for personal use
        "Purchase_Name": "REPLACE",  # For internal and UOSU 
        "Purchase_Date": "Bot_Holder",  # For internal and UOSU 
        "Description": "REPLACE",  # UOSU 
        "Supplier": "Bot_Holder", # UOSU
        "Cost": "Bot_Holder",   # For internal and UOSU 
        "Message": "Bot_Holder",  # For internal 
        "Purchaser": "Bot_Holder", # For internal
        "Receipt_Number": "Bot_Holder",  # For internal and UOSU 
        "Reimbursed": 'No',  # For internal
        "Error_Flag": 'Null' # For internal
    }


    while True: 
        try: 
            response = client.conversations_history(
                channel=chan_id,
                oldest= last_ts,
                cursor = cursor, 
                limit=200,
            )

            messages = response.get('messages',[])
            # Breaks down into individual messages 
            for m in reversed(messages): 
                for f in (m.get("files") or []):
                    all_files.append((m, f))
               
                    
        except SlackApiError as e:
            print("API error:", e.response.get("error"))
            break 

        cursor = response.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
    
    for m,f in all_files: 
        # ["Download_Date", "Purchase_Name", "Description", "Supplier", "Cost", "Message", "Purchaser","Receipt_Number", "Reimbursed"]
        user_name = user_map.get(f["user"])
        file_name = f["name"]
        newest_ts = m.get("ts")
        
        url = f.get("url_private_download")       
        downloaded_path = DOWNLOAD_DIR / file_name
        download_files(url, downloaded_path)

        invoice_num, new_path = change_file_name(downloaded_path, DOWNLOAD_DIR, make_invoice)

        info = {
            "Download_Date": datetime.fromtimestamp(float(m.get("ts"))).strftime("%Y-%m-%d"),
            "Message": m.get("text"),
            "Purchaser": user_name ,
            "Receipt_Number": invoice_num,
            "Reimbursed": "False"
        }

        row = {**defaults, **info}

        rows.append(row)

        with open(JSON_FILE, "w") as f:
            json.dump({"last_ts": newest_ts},f)

        upload_collection_excel_local(row)
        format_excel_output()

if __name__ == "__main__":
    channel_history(CHANNEL_ID)
    format_excel_output()